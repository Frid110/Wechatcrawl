import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from fake_useragent import UserAgent
import tkinter as tk
import openpyxl
#from bs4 import BeautifulSoup

'''def get_content(url):
    resp = urllib.request.urlopen(url)
    h_txt = resp.read()
    bs = BeautifulSoup(h_txt, "html.parser")
    return bs.textarea.get_text()'''



wx_id = 'gh_d7e2f742fc42'

service_args=['--disk-cache=yes', '--ignore-ssl-errors=true','--ssl-protocol=any']
#service_args.append('--load-images=no')  ##关闭图片加载
#service_args.append('--disk-cache=yes')  ##开启缓存
#service_args.append('--ignore-ssl-errors=true') ##忽略https错误  
#service_args.append('--ssl-protocol=any')


cap = webdriver.DesiredCapabilities.PHANTOMJS
cap["phantomjs.page.settings.resourceTimeout"] = 1000
cap["phantomjs.page.settings.loadImages"] = True
cap["phantomjs.page.settings.disk-cache"] = True

ua = UserAgent()
headers = ua.random
cap["phantomjs.page.settings.userAgent"] = headers

driver = webdriver.PhantomJS(executable_path = ".\\phantomjs.exe", desired_capabilities=cap, service_args=service_args) #打开网站
driver.implicitly_wait(10)        ##设置超时时间
driver.set_page_load_timeout(10)  ##设置超时时间

driver.get("https://weixin.sogou.com/", )

try:
        #输入公众号id
        driver.find_element_by_id("query").send_keys(wx_id)
        time.sleep(1)

        #点击搜索
        driver.find_element_by_xpath(r"/html/body/div[3]/div[2]/div/div/form/div/input[4]").click()
        time.sleep(2)

        driver.find_element_by_xpath(r"//*[@uigs='account_article_0']").click()
        time.sleep(2)

        #切换到新窗口
        n = driver.window_handles
        driver.switch_to.window (n[1])
        time.sleep(1)
        '''#模拟滚动
        driver.execute_script('window.document.body.scrollTop = 0')
        driver.execute_script('document.body.scrollTop=document.body.scrollHeight')
        driver.implicitly_wait(20)'''

        #保存网页
        filename = '.\\文章存储位置\\文章导出记录.xlsx'
        url = driver.current_url
        html = driver.page_source.encode('utf-8', "ignore")
        publish_time = driver.find_element_by_id('publish_time').text
        content = driver.find_element_by_id('page-content').text
        try:
                title = driver.find_element_by_id('activity-name').text
                f = open(".\\文章存储位置\\%s.html" %title,'wb')
                f.write(driver.page_source.encode('utf-8', "ignore")) 

                #在Excel表中记录信息
                data = openpyxl.load_workbook(filename)
                table = data.sheetnames[0]#取第一张表
                table = data.active
                mrows = table.max_row
                #信息录入（有题目）
                table.cell(mrows+1,1).value = title
                table.cell(mrows+1,2).value = publish_time
                table.cell(mrows+1,3).value = url
                table.cell(mrows+1,4).value = content
                data.save(filename)   

        except:
                #保存网页文件
                f = open('.\\文章存储位置\\%s.html' %publish_time, 'wb')
                f.write(driver.page_source.encode('utf-8', "ignore"))

                #在Excel表中记录信息
                data = openpyxl.load_workbook(filename)
                table = data.sheetnames[0]#从总表格中取第一张表
                table = data.active
                mrows = table.max_row
                #信息录入（无目录）
                table.cell(mrows+1,2).value = publish_time
                table.cell(mrows+1,3).value = url
                table.cell(mrows+1,4).value = content
                data.save(filename)
        
        window = tk.Tk()
        window.title("提示")
        window.geometry("400x400")
        l = tk.Label(window, text='文章保存成功', font=('Arial', 20), width=30, height=2)
        l.pack()   
        window.mainloop()



except:

        window = tk.Tk()
        window.title("提示")
        window.geometry("400x400")
        l = tk.Label(window, text='文章保存失败', font=('Arial', 20), width=30, height=2)
        l.pack()   
        window.mainloop()
    
driver.quit()

